import os
import re
import json
import shutil
import sys
import subprocess


# ===================== УСТАНОВКА ЗАВИСИМОСТЕЙ =====================

def _ensure_dependencies():
    """Проверяет наличие библиотек. При ошибке импорта устанавливает из requirements.txt и возвращает False (нужен перезапуск)."""
    required = [
        ("openpyxl", "openpyxl"),
        ("colorama", "colorama"),
        ("win32com.client", "pywin32"),
    ]
    for mod_name, _pip_name in required:
        try:
            __import__(mod_name)
        except ImportError:
            break
    else:
        return True

    script_dir = os.path.dirname(os.path.abspath(__file__))
    req_path = os.path.join(script_dir, "requirements.txt")
    if not os.path.isfile(req_path):
        print("Файл requirements.txt не найден. Установите зависимости вручную: pip install -r requirements.txt")
        sys.exit(1)
    print("Установка необходимых библиотек...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", req_path])
    return False


if __name__ == "__main__":
    if not _ensure_dependencies():
        subprocess.run([sys.executable, __file__] + sys.argv[1:])
        sys.exit(0)


from openpyxl import load_workbook
from colorama import Fore, Style, init
import win32com.client as win32

init(autoreset=True)

SPEC_SHEET_NAME = "Specification"
CONFIG_FILE = "config.json"


# ===================== ЦВЕТА =====================

def yellow(text): return Fore.YELLOW + text + Style.RESET_ALL
def green(text): return Fore.GREEN + text + Style.RESET_ALL
def red(text): return Fore.LIGHTRED_EX + text + Style.RESET_ALL


# ===================== КОНФИГ =====================

def script_dir():
    return os.path.dirname(os.path.abspath(__file__))


def load_config():
    path = os.path.join(script_dir(), CONFIG_FILE)
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(data):
    path = os.path.join(script_dir(), CONFIG_FILE)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ===================== ВСПОМОГАТЕЛЬНЫЕ =====================

def normalize_path(path):
    return path.strip().strip('"').strip("'")


def parse_ranges(user_input):
    result = set()
    cleaned = user_input.replace(" ", "").replace(";", ",")
    for part in cleaned.split(","):
        if "-" in part:
            a, b = part.split("-")
            result.update(range(int(a), int(b) + 1))
        elif part:
            result.add(int(part))
    return result


def extract_invoice_number(folder):
    m = re.match(r"(\d+)", folder)
    return int(m.group(1)) if m else None


def hide_all_except_spec_xlsx(path):
    wb = load_workbook(path)
    if SPEC_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Лист "{SPEC_SHEET_NAME}" не найден')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == SPEC_SHEET_NAME:
            ws.sheet_state = "visible"
        else:
            ws.sheet_state = "hidden"

    wb.save(path)


# ===================== СЦЕНАРИИ =====================

def scenario_xlsx(folder, number):
    src = os.path.join(folder, f"Invoice {number}.xlsx")
    dst = os.path.join(folder, f"Invoice {number} fcs.xlsx")
    shutil.copy2(src, dst)
    hide_all_except_spec_xlsx(dst)


def scenario_xls(folder, number):
    src = os.path.join(folder, f"Invoice {number}.xlsx")
    dst = os.path.join(folder, f"Invoice {number} fcs.xls")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(src)
        wb.SaveAs(dst, FileFormat=56)
        wb.Close(SaveChanges=False)

        wb = excel.Workbooks.Open(dst)
        for sheet in wb.Worksheets:
            sheet.Visible = sheet.Name == SPEC_SHEET_NAME
        wb.Save()
        wb.Close()
    finally:
        excel.Quit()


def scenario_delete_fcs(folder, number):
    deleted = 0

    for fname in os.listdir(folder):
        name, ext = os.path.splitext(fname)

        if ext.lower() not in {".xlsx", ".xls"}:
            continue

        if not name.lower().endswith("fcs"):
            continue

        full_path = os.path.join(folder, fname)

        if os.path.isfile(full_path):
            os.remove(full_path)
            deleted += 1
            print(green(f"[УДАЛЕНО] {fname}"))

    if deleted == 0:
        print(yellow(f"[НЕТ ФАЙЛОВ] Invoice {number} — fcs файлы не найдены"))

    return deleted


# ===================== MAIN =====================

def main():
    config = load_config()
    base_dir = config.get("base_dir")

    while True:
        if not base_dir:
            print(yellow("Укажите путь к директории с инвойсами:"))
            base_dir = normalize_path(input("> "))
            if not os.path.isdir(base_dir):
                print(red("Некорректный путь"))
                base_dir = None
                continue
            save_config({"base_dir": base_dir})

        print(yellow("\nВыберите сценарий:"))
        print(yellow("1) Создание файла спецификации XLS") + green("X"))
        print(yellow("2) Создание файла спецификации XLS"))
        current_path = base_dir if base_dir else "не указан"
        print(yellow(f'3) Изменить путь к директории с инвойсами (Сейчас: {current_path})'))
        print(yellow("4) Удаление файлов спецификаций (Invoice ... fcs)"))
        print(yellow("0) Выход"))

        choice = input("> ").strip()

        if choice == "0":
            break

        if choice == "3":
            base_dir = None
            continue

        if choice not in {"1", "2", "4"}:
            print(red("Неверный выбор"))
            continue

        while True:
            if choice == "4":
                print(green("\nНажмите Enter ") + yellow("или укажите диапазон номеров:"))
            else:
                print(yellow("\nУкажите диапазон номеров инвойсов:"))
            user_input = input("> ").strip()

            if choice == "4" and user_input == "":
                matched = []
                for folder in os.listdir(base_dir):
                    num = extract_invoice_number(folder)
                    if num is not None:
                        matched.append((num, os.path.join(base_dir, folder)))
            else:
                invoice_numbers = parse_ranges(user_input)
                matched = []
                for folder in os.listdir(base_dir):
                    num = extract_invoice_number(folder)
                    if num in invoice_numbers:
                        matched.append((num, os.path.join(base_dir, folder)))

            if not matched:
                print(red("Инвойсы указанных номеров не были найдены"))
                continue

            processed = 0
            for number, folder in sorted(matched):
                print(yellow(f"[В РАБОТЕ] Invoice {number}"))
                try:
                    if choice == "1":
                        scenario_xlsx(folder, number)
                    elif choice == "2":
                        scenario_xls(folder, number)
                    elif choice == "4":
                        scenario_delete_fcs(folder, number)

                    processed += 1
                    print(green(f"[ГОТОВО] Invoice {number}"))
                except Exception as e:
                    print(red(f"[ОШИБКА] Invoice {number}: {e}"))

            if choice == "4":
                print(green("\nГотово! Удаление завершено"))
            else:
                print(green(f"\nГотово! Обработано файлов: {processed}"))
            break

    print(green("\nВыход из утилиты"))
    input(yellow("Нажмите Enter для закрытия окна"))


if __name__ == "__main__":
    main()
